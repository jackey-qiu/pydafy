import os
import copy
import numpy as np
import pandas as pd
import logging
from PyQt5.QtWidgets import QFileDialog, QInputDialog, QCheckBox, QTableWidgetItem
from PyQt5.QtGui import QFont, QBrush, QColor
from dafy.core.FilterPool import data
from dafy.core.util.UtilityFunctions import locate_tag, apply_modification_of_code_block as script_block_modifier
from dafy.projects.superrod.core.models.structure_tools import sorbate_tool
from dafy.core.util.path import user_data_path, user_example_path

class SuperRodIoControl(object):
    def load_rod_files(self):
        '''
        load all rod files(*.rod) located in a selected folder
        '''
        folder = str(QFileDialog.getExistingDirectory(self, "Select Directory"))
        self.lineEdit_folder_of_rod_files.setText(folder)
        for file in os.listdir(folder):
            if file.endswith('.rod'):
                self.listWidget_rod_files.addItem(file)

    def remove_rod_files(self):
        '''
        remove all files in the list
        '''
        self.listWidget_rod_files.clear()

    def remove_selected_rod_files(self):
        items = self.listWidget_rod_files.selectedItems()
        if not items:return
        for item in items:
            self.listWidget_rod_files.takeItem(self.listWidget_rod_files.row(item))

    def load_next_rod_file_in_batch(self):
        if not hasattr(self.run_batch,'rod_files'):
            pass
        else:
            #if you are now at the last rod file, then roll back to the first one
            if self.run_batch.rod_files.index(self.rod_file)==(len(self.run_batch.rod_files)-1):
                self.open_model_with_path(self.run_batch.rod_files[0])
                self.listWidget_rod_files.setCurrentRow(0)
            else:
                target_index = self.run_batch.rod_files.index(self.rod_file) + 1
                self.open_model_with_path(self.run_batch.rod_files[target_index])
                self.listWidget_rod_files.setCurrentRow(target_index)

    def load_previous_rod_file_in_batch(self):
        if not hasattr(self.run_batch,'rod_files'):
            pass
        else:
            #if you are now at the first rod file, then roll forward to the last one
            if self.run_batch.rod_files.index(self.rod_file)==0:
                self.open_model_with_path(self.run_batch.rod_files[-1])
                self.listWidget_rod_files.setCurrentRow(len(self.run_batch.rod_files)-1)
            else:
                target_index = self.run_batch.rod_files.index(self.rod_file) - 1
                self.open_model_with_path(self.run_batch.rod_files[target_index])
                self.listWidget_rod_files.setCurrentRow(target_index)

    def open_model_with_path(self,path):
        fileName = path
        load_add_ = 'success'
        self.rod_file = fileName
        if fileName:
            try:
                self.setWindowTitle('Data analysis factory: CTR data modeling-->{}'.format(fileName))
                self.model.load(fileName)
                # self.load_addition()
                try:
                    self.load_addition()
                except:
                    load_add_ = 'failure'
                #add a mask attribute to each dataset
                for each in self.model.data_original:
                    if not hasattr(each,'mask'):
                        each.mask = np.array([True]*len(each.x))
                for each in self.model.data:
                    if not hasattr(each,'mask'):
                        each.mask = np.array([True]*len(each.x))
                #add model space to terminal
                self.widget_terminal.update_name_space("model",self.model)
                self.widget_terminal.update_name_space("solver",self.run_fit.solver)
                self.widget_terminal.update_name_space("win",self)

                #remove items in the msv and re-initialize it
                self.widget_edp.items = []
                # self.widget_msv_top.items = []
                #update other pars
                self.update_table_widget_data()
                self.update_plot_dimension()
                self.update_combo_box_dataset()
                self.update_plot_data_view()
                self.update_par_upon_load()
                self.update_script_upon_load()
                #model is simulated at the end of next step
                self.init_mask_info_in_data_upon_loading_model()
                #add name space for cal bond distance after simulation
                try:
                    self.widget_terminal.update_name_space("report_distance",self.model.script_module.sample.inter_atom_distance_report)
                except:
                    pass
                #now set the comboBox for par set
                self.update_combo_box_list_par_set()

                self.statusbar.clearMessage()
                self.statusbar.showMessage("Model is loaded, and {} in config loading".format(load_add_))
                # self.update_mask_info_in_data()
            except Exception:

                self.statusbar.clearMessage()
                self.statusbar.showMessage('Failure to open a model file!')
                logging.root.exception('Fatal error encountered during openning a model file!')
                self.tabWidget_data.setCurrentIndex(7)

    def open_model(self):
        """open a saved model file(*.rod), which is a compressed file containing data, script and fit parameters in one place"""
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", str(user_example_path),"rod file (*.rod);;zip Files (*.rar)", options=options)
        self.open_model_with_path(fileName)
        self.lineEdit_folder_of_rod_files.setText(os.path.dirname(fileName))
        self.listWidget_rod_files.addItem(os.path.basename(fileName))

    def open_model_selected_in_listWidget(self):
        """
        open a saved model file(*.rod), which is a compressed file containing data, script and fit parameters in one place
        The rod file the the selected rod file in the listWidget
        """
        fileName = os.path.join(self.lineEdit_folder_of_rod_files.text(),self.listWidget_rod_files.currentItem().text())
        self.open_model_with_path(fileName)

    def auto_save_model(self):
        """model will be saved automatically during fit, for which you need to set the interval generations for saving automatically"""
        #the model will be renamed this way
        self.update_par_upon_change()
        path = self.rod_file.replace(".rod","_ran.rod")
        if path:
            #update the error bar
            self.calculate_error_bars()
            self.model.script = (self.plainTextEdit_script.toPlainText())
            # self.model.save(path)
            save_add_ = 'success'
            try:
                self.model.save_all(path, self.run_fit.solver.optimizer)
                # self.save_addition()
            except:
                save_add_ = "failure"
            self.statusbar.clearMessage()
            self.statusbar.showMessage("Model is saved, and {} in config saving".format(save_add_))

    def save_model(self):
        """model will be saved automatically during fit, for which you need to set the interval generations for saving automatically"""
        #the model will be renamed this way
        try:
            path = self.rod_file
            try:
                self.calculate_error_bars()
            except:
                pass
            #self.model.script = (self.plainTextEdit_script.toPlainText())
            self.update_data_check_attr()
            self.update_par_upon_change()
            self.model.script = (self.plainTextEdit_script.toPlainText())
            self.widget_solver.update_parameter_in_solver(self)
            self.model.save(path)
            save_add_ = 'success'
            try:
                self.save_addition()
            except:
                save_add_ = "failure"
            self.statusbar.clearMessage()
            self.statusbar.showMessage("Model is saved, and {} in config saving".format(save_add_))
        except Exception:
            self.statusbar.clearMessage()
            self.statusbar.showMessage('Failure to save model!')
            logging.root.exception('Fatal error encountered during model save!')
            self.tabWidget_data.setCurrentIndex(7)

    def save_model_as(self):
        """save model file, promting a dialog widget to ask the file name to save model"""
        path, _ = QFileDialog.getSaveFileName(self, "Save file as", "", "rod file (*.rod);;zip files (*.rar)")
        if path:
            #update the rod_file attribute
            self.rod_file = path
            try:
                self.calculate_error_bars()
            except:
                pass
            #self.model.script = (self.plainTextEdit_script.toPlainText())
            self.update_data_check_attr()
            self.update_par_upon_change()
            self.model.script = (self.plainTextEdit_script.toPlainText())
            self.widget_solver.update_parameter_in_solver(self)
            self.model.save(path)
            save_add_ = 'success'
            try:
                self.save_addition()
            except:
                save_add_ = "failure"
            self.statusbar.clearMessage()
            self.statusbar.showMessage("Model is saved, and {} in config saving".format(save_add_))
            self.setWindowTitle('Data analysis factory: CTR data modeling-->{}'.format(path))

    #here save also the config pars for diffev solver
    def save_addition(self):
        """save solver parameters, pulling from pyqtgraphy.parameter_tree widget"""
        values=\
                [self.widget_solver.par.param('Diff.Ev.').param('k_m').value(),
                self.widget_solver.par.param('Diff.Ev.').param('k_r').value(),
                self.widget_solver.par.param('Diff.Ev.').param('Method').value(),
                self.widget_solver.par.param('FOM').param('Figure of merit').value(),
                self.widget_solver.par.param('FOM').param('Auto save, interval').value(),
                self.widget_solver.par.param('FOM').param('weighting factor').value(),
                self.widget_solver.par.param('FOM').param('weighting region').value(),
                self.widget_solver.par.param('Fitting').param('start guess').value(),
                self.widget_solver.par.param('Fitting').param('Generation size').value(),
                self.widget_solver.par.param('Fitting').param('Population size').value()]
        pars = ['k_m','k_r','Method','Figure of merit','Auto save, interval','weighting factor','weighting region','start guess','Generation size','Population size']
        for i in range(len(pars)):
            self.model.save_addition(pars[i],str(values[i]))
        model_info = ''
        if hasattr(self,'textEdit_note'):
            model_info = self.textEdit_note.toPlainText()
        self.model.save_addition('model_info',model_info)
        # print(str(self.textEdit_cov.toHtml()))
        if hasattr(self, 'covariance_matrix'):
            # self.model.save_addition('covariance_matrix',str(self.textEdit_cov.toHtml()))
            self.model.save_addition('covariance_matrix',self.covariance_matrix)
        else:
            self.model.save_addition('covariance_matrix',pd.DataFrame(np.identity(10)))
        if hasattr(self, 'sensitivity_data'):
            self.model.save_addition('sensitivity',str(self.sensitivity_data))
            # print(str(self.sensitivity_data))
        else:
            self.model.save_addition('sensitivity',str([]))
            # print(pars[i],str(values[i]))

    def save_log_info(self):
        from datetime import datetime
        with open(user_data_path / f'LOG_INFO_{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.txt', 'w') as f:
            f.write(self.textBrowser_error_msg.document().toPlainText())
            logging.root.info('Log info is saved in the dump_files!')

    def load_addition(self):
            funcs=\
                [self.widget_solver.par.param('Diff.Ev.').param('k_m').setValue,
                self.widget_solver.par.param('Diff.Ev.').param('k_r').setValue,
                self.widget_solver.par.param('Diff.Ev.').param('Method').setValue,
                self.widget_solver.par.param('FOM').param('Figure of merit').setValue,
                self.widget_solver.par.param('FOM').param('Auto save, interval').setValue,
                self.widget_solver.par.param('FOM').param('weighting factor').setValue,
                self.widget_solver.par.param('FOM').param('weighting region').setValue,
                self.widget_solver.par.param('Fitting').param('start guess').setValue,
                self.widget_solver.par.param('Fitting').param('Generation size').setValue,
                self.widget_solver.par.param('Fitting').param('Population size').setValue]

            types= [float,float,str,str,int,float,str,bool,int,int]
            pars = ['k_m','k_r','Method','Figure of merit','Auto save, interval','weighting factor','weighting region','start guess','Generation size','Population size']
            value = None
            for i in range(len(pars)):
                type_ = types[i]
                if type_ == float:
                    try:
                        value = np.round(float(self.model.load_addition(pars[i])),2)
                    except:
                        pass
                elif type_==str:
                    try:
                        value = self.model.load_addition(pars[i]).decode("utf-8")
                    except:
                        pass
                elif type_==bool:
                    try:
                        value = (self.model.load_addition(pars[i]).decode("ASCII")=="True")
                    except:
                        pass
                else:
                    try:
                        value = type_(self.model.load_addition(pars[i]))
                    except:
                        pass
                if value!=None:
                    funcs[i](value)
            model_info = ''
            sensitivity_data = []
            covariance_matrix = pd.DataFrame(np.identity(3))
            try:
                model_info = self.model.load_addition('model_info').decode('utf-8')
            except:
                pass
            try:
                sensitivity_data = eval(self.model.load_addition('sensitivity').decode('utf-8'))
            except:
                pass
            try:
                covariance_matrix = self.model.load_addition('covariance_matrix', load_type = 'object')
            except:
                pass
            if hasattr(self,'textEdit_note'):
                self.textEdit_note.setPlainText(model_info)
            # self.textEdit_cov.setHtml(covariant_matrix)
            self.sensitivity_data = sensitivity_data
            self.covariance_matrix = covariance_matrix
            self.textEdit_cov.setHtml(covariance_matrix.style.background_gradient(cmap='coolwarm').set_precision(3).render())            
            self.plot_bar_chart(sensitivity_data)

    def load_data(self, loader = 'ctr'):
        self._empty_data_pool()
        exec('self.load_data_{}()'.format(loader))

    def append_data(self):
        self.load_data_ctr()

    def _empty_data_pool(self):
        #now empty the data pool
        self.model.data.items = [data.DataSet(name='Data 0')]
        self.model.data._counter = 1

    def load_data_ctr(self):
        """
        load data
        ------------
        if the data is ctr data, then you should stick to the dataformat as follows
        #8 columns in total
        #X, H, K, Y, I, eI, LB, dL
        #for CTR data, X column is L column, Y column all 0
        #for RAXR data, X column is energy column, Y column is L column
        #H, K, columns are holding H, K values
        #I column is holding background-subtraced intensity of ctr signal
        #LB, and dL are two values for roughness calculation
           LB: first Bragg peak L of one rod
           dL: interval L between two adjacent Bragg peak L's
        To get access to these columns:
            X column: data.x
            I column: data.y
            eI column: data.error
            H column: data.extra_data["h"]
            K column: data.extra_data["k"]
            Y column: data.extra_data["Y"]
            LB column: data.extra_data["LB"]
            dL column: data.extra_data["dL"]
        ---------------
        if the data you want to load is not in CTR format, to make successful loading, assure:
            1)your data file has 8 columns
            2)columns are space-separated (or tab-seperated)
            3)you can add comment lines heading with "#"
            4)if your data has <8 columns, then fill the other unused columns with 0
            5)to asscess your data column, you should use the naming rule described above, although
              the real meaning of each column, eg X column, could be arbitrary at your wishes
              For example, you put frequence values to the first column(X column), then to access this
              column, you use data.X

        Data file of 8 columns should be enough to encountpass many different situations.
        """
        self.model.compiled = False
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","csv Files (*.csv);;data Files (*.dat);txt Files (*.txt)", options=options)
        current_data_set_name = [self.tableWidget_data.item(i,0).text() for i in range(self.tableWidget_data.rowCount())]
        if fileName:
            with open(fileName,'r') as f:
                data_loaded = np.loadtxt(f,comments = '#',delimiter=None)
                data_loaded_pd = pd.DataFrame(data_loaded, columns = ['X','h','k','Y','I','eI','LB','dL'])
                data_loaded_pd['h'] = data_loaded_pd['h'].apply(lambda x:int(np.round(x)))
                data_loaded_pd['k'] = data_loaded_pd['k'].apply(lambda x:int(np.round(x)))
                data_loaded_pd.sort_values(by = ['h','k','Y'], inplace = True)
                hk_unique = list(set(zip(list(data_loaded_pd['h']), list(data_loaded_pd['k']), list(data_loaded_pd['Y']))))
                hk_unique.sort()
                h_unique = [each[0] for each in hk_unique]
                k_unique = [each[1] for each in hk_unique]
                Y_unique = [each[2] for each in hk_unique]

                for i in range(len(h_unique)):
                    h_temp, k_temp, Y_temp = h_unique[i], k_unique[i], Y_unique[i]
                    if Y_temp==0:#CTR data
                        name = 'Data-{}{}L'.format(h_temp, k_temp)
                    else:#RAXR data
                        name = 'Data-{}{}_L={}'.format(h_temp, k_temp, Y_temp)
                    tag = sum([int(name in each) for each in current_data_set_name])+1
                    #if name in current_data_set_name:
                    name = name + '_{}'.format(tag)
                    self.model.data.add_new(name = name)
                    sub_data = data_loaded_pd[(data_loaded_pd['h']==h_temp) & (data_loaded_pd['k']==k_temp)& (data_loaded_pd['Y']==Y_temp)]
                    sub_data.sort_values(by='X',inplace =True)
                    self.model.data.items[-1].x = sub_data['X'].to_numpy()
                    self.model.data.items[-1].y = sub_data['I'].to_numpy()
                    self.model.data.items[-1].error = sub_data['eI'].to_numpy()
                    self.model.data.items[-1].x_raw = sub_data['X'].to_numpy()
                    self.model.data.items[-1].y_raw = sub_data['I'].to_numpy()
                    self.model.data.items[-1].error_raw = sub_data['eI'].to_numpy()
                    self.model.data.items[-1].set_extra_data(name = 'h', value = sub_data['h'].to_numpy())
                    self.model.data.items[-1].set_extra_data(name = 'k', value = sub_data['k'].to_numpy())
                    self.model.data.items[-1].set_extra_data(name = 'Y', value = sub_data['Y'].to_numpy())
                    self.model.data.items[-1].set_extra_data(name = 'LB', value = sub_data['LB'].to_numpy())
                    self.model.data.items[-1].set_extra_data(name = 'dL', value = sub_data['dL'].to_numpy())
                    self.model.data.items[-1].mask = np.array([True]*len(self.model.data.items[-1].x))
                    # self.model.data.concatenate_all_ctr_datasets()
        #now remove the empty datasets
        empty_data_index = []
        i=0
        for each in self.model.data.items:
            if len(each.x_raw) == 0:
                empty_data_index.append(i)
            i += 1
        for i in range(len(empty_data_index)):
            self.model.data.delete_item(empty_data_index[i])
            for ii in range(len(empty_data_index)):
                if empty_data_index[ii]>empty_data_index[i]:
                    empty_data_index[ii] = empty_data_index[ii]-1
                else:
                    pass
        self.model.data_original = copy.deepcopy(self.model.data)
        self.model.data.concatenate_all_ctr_datasets()
        #update the view
        self.update_table_widget_data()
        self.update_combo_box_dataset()
        self.update_plot_dimension()
        self.update_plot_data_view()

    def delete_data(self):
        self.model.compiled = False
        # Delete the selected mytable lines
        row_index = [each.row() for each in self.tableWidget_data.selectionModel().selectedRows()]
        row_index = sorted(row_index, reverse=True)
        for each in row_index:
            self.model.data.delete_item(each)
            self.model.data_original.delete_item(each)
        self.update_table_widget_data()
        self.update_combo_box_dataset()
        self.update_plot_dimension()
        self.update_plot_data_view()

    def save_structure_file(self):
        domain_tag, done = QInputDialog.getInt(self, 'Domain tag', 'Enter the domain index for the structure you want to save eg 0:')
        if not done:
            domain_tag = 0
        path, _ = QFileDialog.getSaveFileName(self, "Save file", "", "xyz file (*.xyz)")
        if path:
            self.model.script_module.sample.make_xyz_file(which_domain = int(domain_tag), save_file = path)
            self.statusbar.clearMessage()
            self.statusbar.showMessage('The data file is saved at {}'.format(path))

    def save_data(self):
        def _make_data():
            """[append missing points near Bragg peaks, the values for data column at these points will be set to nan, while the values for model column will be calculated]
            """
            extended_data = {}
            keys = ['potential','L', 'H', 'K', 'I', 'I_model', 'error', 'I_bulk', 'use']
            for each in keys:
                extended_data[each] = []
            for data in self.model.data:
                L = data.x
                H = data.extra_data['h']
                K = data.extra_data['k']
                dL = data.extra_data['dL']
                LB = data.extra_data['LB']
                I = data.y
                #I_model = data.y_sim
                error = data.error
                Bragg_L = LB[0] + np.array(range(-2,10))*dL[0]
                Bragg_L = [each for each in Bragg_L if L.max()>each>L.min()]
                Bragg_index = []
                for each_bragg_L in Bragg_L:
                    ix = np.argpartition(abs(L - each_bragg_L),1)
                    left, right = None, None
                    ix_left, ix_right = None, None
                    if L[ix[0]]>each_bragg_L:
                        right = L[ix[0]]
                        ix_right = ix[0]
                        left = L[ix[0]-1]
                        ix_left = ix_right -1
                    else:
                        left = L[ix[0]]
                        ix_left = ix[0]
                        right = L[ix[0]+1]
                        ix_right = ix[0]+1
                    Bragg_index.append([ix_left+num_points_near_Bragg_peak, ix_left+num_points_near_Bragg_peak+1])
                    appended_Ls = list(np.linspace(left, each_bragg_L-0.02, num_points_near_Bragg_peak, endpoint = True))+ list(np.linspace(right, each_bragg_L+0.02, num_points_near_Bragg_peak, endpoint = True))[::-1]
                    appended_Hs = [H[0]]*len(appended_Ls)
                    appended_Ks = [K[0]]*len(appended_Ls)
                    appended_dL = [dL[ix_right]]*len(appended_Ls)
                    appended_LB = [LB[ix_right]]*len(appended_Ls)
                    L = np.concatenate((L[:ix_right],appended_Ls,L[ix_right:]))
                    H = np.concatenate((H[:ix_right],[H[0]]*len(appended_Ls),H[ix_right:]))
                    K = np.concatenate((K[:ix_right],[K[0]]*len(appended_Ls),K[ix_right:]))
                    dL = np.concatenate((dL[:ix_right],[dL[ix_right]]*len(appended_Ls),dL[ix_right:]))
                    LB = np.concatenate((LB[:ix_right],[LB[ix_right]]*len(appended_Ls),LB[ix_right:]))
                    I = np.concatenate((I[:ix_right],[np.nan]*len(appended_Ls),I[ix_right:]))
                    #I_model = np.concatenate((I_model[:ix_right],[np.nan]*len(appended_Ls),I_model[ix_right:]))
                    error = np.concatenate((error[:ix_right],[np.nan]*len(appended_Ls),error[ix_right:]))
                beta = self.model.script_module.rgh.beta
                rough = (1-beta)/((1-beta)**2 + 4*beta*np.sin(np.pi*(L-LB)/dL)**2)**0.5
                f = rough*self.model.script_module.sample.calc_f_all(H, K, L)
                f_ideal = self.model.script_module.sample.calc_f_ideal(H, K, L)
                extra_scale_factor = 'scale_factor_{}{}L'.format(int(round(H[0],0)),int(round(K[0],0)))
                if hasattr(self.model.script_module.rgh,extra_scale_factor):
                    rod_factor = getattr(self.model.script_module.rgh, extra_scale_factor)
                else:
                    if int(round(H[0],0))==0 and int(round(K[0],0))==0:#specular rod
                        if hasattr(self.model.script_module.rgh,'scale_specular_rod'):
                            rod_factor = getattr(self.model.script_module.rgh,'scale_specular_rod')
                        else:
                            rod_factor = 1
                    else:#nonspecular rod
                        if hasattr(self.model.script_module.rgh,'scale_nonspecular_rod'):
                            rod_factor = getattr(self.model.script_module.rgh,'scale_nonspecular_rod')
                        else:
                            rod_factor = 1
                I_model = list(abs(f*f)*self.model.script_module.rgh.scale_nonspecular_rods*rod_factor)
                I_bulk = list(abs(f_ideal*f_ideal)*self.model.script_module.rgh.scale_nonspecular_rods*rod_factor)
                E = [potential]*len(L)
                use = [True]*len(L)
                for each in keys:
                    if each=='potential':
                        new = locals()['E']
                    else:
                        new = locals()[each]
                    extended_data[each] = list(extended_data[each]) + list(new)
            return extended_data

        num_points_near_Bragg_peak = 4
        potential, done = QInputDialog.getDouble(self, 'Potential_info', 'Enter the potential for this dataset (in V):', value = 0.0)
        if not done:
            potential = None
        #for unknown reseason, this line cause windows fatal exception in terminal, but does not stop the main program
        #don't worry about it for now
        path, _ = QFileDialog.getSaveFileName(self, "Save file", "", "data file (*.csv)")
        if path!="":
            export_data = _make_data()
            df_export_data = pd.DataFrame(export_data)
            #some problem with saving the full dataset into xlsx file format due to version incompatibility of openpyxl
            #self._append_df_to_excel(filename = [path+'.xlsx',path][int(path.endswith('.xlsx'))], df = df_export_data, sheet_name = 'Sheet1', startrow = None, truncate_sheet=False, columns = ['potential','L', 'H', 'K', 'I', 'I_model', 'error', 'I_bulk', 'use'])
            
            #file ended with rod could be loaded into superrod program
            #file ended with full could be used to generate graphs using user scripts, this file contains more info the rod data.

            if path.endswith('.csv'):
                path_data_rod = path[0:-4] + '_rod'
                path_data_model = path[0:-4] + '_full.csv'
            else:
                path_data_rod = path + '_rod'
                path_data_model = path + '_full.csv'

            self.save_data_original(path=path_data_rod)
            
            #also save loadable csv file
            df_export_data.to_csv(path_data_model,sep="\t",columns=['potential','L', 'H', 'K', 'I', 'I_model', 'error', 'I_bulk', 'use'],\
                                 index=False, header=['#potential','L', 'H', 'K', 'I', 'I_model', 'error', 'I_bulk', 'use'])
            

    def _append_df_to_excel(self, filename, df, sheet_name='Sheet1', startrow=None,
                        truncate_sheet=False, 
                        **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
        filename : File path or existing ExcelWriter
                    (Example: '/path/to/file.xlsx')
        df : dataframe to save to workbook
        sheet_name : Name of sheet which will contain DataFrame.
                    (default: 'Sheet1')
        startrow : upper left cell row to dump data frame.
                    Per default (startrow=None) calculate the last row
                    in the existing DF and write to the next row...
        truncate_sheet : truncate (remove and recreate) [sheet_name]
                        before writing DataFrame to Excel file
        to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """
        from openpyxl import load_workbook

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, mode='rb', engine='openpyxl')
        header = None

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
            header = False
        except FileNotFoundError:
            # file does not exist yet, we will create it
            header = True
            #pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, header = header, **to_excel_kwargs)

        # save the workbook
        writer.save()

    #save data plus best fit profile
    def save_data_original(self, path=""):
        '''
        potential, done = QInputDialog.getDouble(self, 'Potential_info', 'Enter the potential for this dataset (in V):')
        if not done:
            potential = None
        path, _ = QFileDialog.getSaveFileName(self, "Save file", "", "data file (*.*)")
        '''
        if path!="":
            keys_attri = ['x','y','y_sim','error']
            keys_extra = ['h','k','Y','dL','LB']
            lib_map = {'x': 'L', 'y':'I','y_sim':'I_model','error':'error','h':'H','k':'K','Y':'Y','dL':'dL','LB':'LB'}
            export_data = {}
            for key in ['x','h','k','y','y_sim','error','Y','dL','LB']:
                export_data[lib_map[key]] = []
            export_data['use'] = []
            #export_data['I_bulk'] = []
            #export_data['potential'] = []
            for each in self.model.data:
                if each.use:
                    for key in ['x','h','k','y','y_sim','error','Y','dL','LB']:
                        if key in keys_attri:
                            export_data[lib_map[key]] = np.append(export_data[lib_map[key]], getattr(each,key))
                        elif key in keys_extra:
                            export_data[lib_map[key]] = np.append(export_data[lib_map[key]], each.extra_data[key])
                    export_data['use'] = np.append(export_data['use'],[True]*len(each.x))
                else:
                    for key in ['x','h','k','y','y_sim','error','Y','dL','LB']:
                        if key in keys_attri:
                            if key=='y_sim':
                                export_data[lib_map[key]] = np.append(export_data[lib_map[key]], [0]*len(getattr(each,'x')))
                            else:
                                export_data[lib_map[key]] = np.append(export_data[lib_map[key]], getattr(each,key))
                        elif key in keys_extra:
                            export_data[lib_map[key]] = np.append(export_data[lib_map[key]], each.extra_data[key])
                    export_data['use'] = np.append(export_data['use'],[False]*len(each.x))
                '''
                export_data['potential'] = np.append(export_data['potential'],[float(potential)]*len(each.x))
                beta = self.model.script_module.rgh.beta
                #rough = (1-beta)/((1-beta)**2 + 4*beta*np.sin(np.pi*(each.x-each.extra_data['LB'])/each.extra_data['dL'])**2)**0.5
                scale_factor = [self.model.script_module.rgh.scale_nonspecular_rods,self.model.script_module.rgh.scale_specular_rod][int("00L" in each.name)]
                h_, k_ = int(round(each.extra_data['h'][0],0)),int(round(each.extra_data['k'][0],0))
                extra_scale_factor = 'scale_factor_{}{}L'.format(h_,k_)
                if hasattr(self.model.script_module.rgh,extra_scale_factor):
                    rod_factor = getattr(self.model.script_module.rgh, extra_scale_factor)
                else:
                    rod_factor = 1
                rough = 1
                export_data['I_bulk'] = np.append(export_data['I_bulk'],rough**2*np.array(self.model.script_module.sample.calc_f_ideal(each.extra_data['h'], each.extra_data['k'], each.x)**2*scale_factor*rod_factor))
                '''
            '''
            writer_temp = pd.ExcelWriter([path+'.xlsx',path][int(path.endswith('.xlsx'))])
            df_export_data.to_excel(writer_temp, columns =['potential']+[lib_map[each_] for each_ in ['x','h','k','y','y_sim','error']]+['I_bulk','use'])
            writer_temp.save()
            writer_temp.close()
            '''
            #also save loadable csv file
            df_export_data = pd.DataFrame(export_data)
            df_export_data.to_csv([path+'.csv',path][int(path.endswith('.csv'))],sep="\t",columns=['L','H','K','Y','I','error','LB','dL'],\
                                 index=False, header=['#L','H','K','Y','I','error','LB','dL'])

    def load_script(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","script Files (*.py);;text Files (*.txt)", options=options)
        if fileName:
            with open(fileName,'r') as f:
                self.plainTextEdit_script.setPlainText(f.read())
        self.model.script = (self.plainTextEdit_script.toPlainText())

    def update_script_upon_load(self):
        self.plainTextEdit_script.setPlainText(self.model.script)

    def save_script(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save script file", "", "script file (*.py)")
        with open(path,'w') as f:
            f.write(self.model.script)

    def modify_script(self):
        """
        Modify script based on the specified sorbates and total domain numbers
        To use this function, your script file should be standadized to have 
        tags to specifpy the code block position where you define the sorbates.
        This func is customized to modify script_model_standard.py.
        """
        assert self.model.script!="","No script to work on, please load script first!"
        domain_num = int(self.lineEdit_domain_number.text().rstrip())
        motif_chain = self.lineEdit_sorbate_motif.text().strip().rsplit(",")

        assert domain_num == len(motif_chain), "Number of domain not match with the motif number. Fix it first!!"
        lines = script_block_modifier(self.model.script.rsplit("\n"), 'slabnumber',["num_surface_slabs"],[domain_num])

        els_sorbate = []
        anchor_index_list = []
        flat_down_index = []
        xyzu_oc_m = []
        structure = []
        for each in motif_chain:
            each = each.strip()
            properties_temp = getattr(sorbate_tool,each)
            for each_key in properties_temp:
                if each_key == "els_sorbate":
                    els_sorbate.append(properties_temp[each_key])
                elif each_key == "anchor_index_list":
                    anchor_index_list.append(properties_temp[each_key])
                elif each_key == "flat_down_index":
                    flat_down_index.append(properties_temp[each_key])
                elif each_key == "structure":
                    structure.append("#"+each+properties_temp[each_key])
        xyzu_oc_m = [[0.5, 0.5, 1.5, 0.1, 1, 1]]*len(els_sorbate)
        tag_list = ['els_sorbate', 'anchor_index_list', 'flat_down_index', 'xyzu_oc_m']
        tag_value_list = [els_sorbate, anchor_index_list, flat_down_index, xyzu_oc_m]
        lines = script_block_modifier(lines, 'sorbateproperties',tag_list, tag_value_list)
        left_, right_ = locate_tag(lines,'sorbatestructure')
        del(lines[left_:right_])
        if structure[-1][-1] == "\n":
            structure[-1] = structure[-1][0:-1]
        lines.insert(left_,"\n".join(structure))

        self.model.script = '\n'.join(lines)
        self.plainTextEdit_script.setPlainText(self.model.script)

    def _load_par(self):
        vertical_labels = []
        self.tableWidget_pars.setRowCount(1)
        self.tableWidget_pars.setColumnCount(7)
        self.tableWidget_pars.setHorizontalHeaderLabels(['Parameter','Value','Fit','Min','Max','Error','Link'])
        items = ['par',0,'False',0,0,'-','']
        for i in [0]:
            j = 0
            if items[0] == '':
                self.model.parameters.data.append([items[0],0,False,0, 0,'-',''])
                vertical_labels.append('')
                j += 1
            else:
                #add items to parameter attr
                self.model.parameters.data.append([items[0],float(items[1]),items[2]=='True',float(items[3]), float(items[4]),items[5],items[6]])
                #add items to table view
                if len(vertical_labels)==0:
                    vertical_labels.append('1')
                else:
                    if vertical_labels[-1] != '':
                        vertical_labels.append('{}'.format(int(vertical_labels[-1])+1))
                    else:
                        vertical_labels.append('{}'.format(int(vertical_labels[-2])+1))
                for item in items:
                    if j == 2:
                        check_box = QCheckBox()
                        check_box.setChecked(item=='True')
                        self.tableWidget_pars.setCellWidget(i,2,check_box)
                    else:
                        qtablewidget = QTableWidgetItem(item)
                        # qtablewidget.setTextAlignment(Qt.AlignCenter)
                        if j == 0:
                            qtablewidget.setFont(QFont('Times',10,QFont.Bold))
                        elif j == 1:
                            qtablewidget.setForeground(QBrush(QColor(255,0,255)))
                        self.tableWidget_pars.setItem(i,j,qtablewidget)
                    j += 1
        # self.tableWidget_pars.resizeColumnsToContents()
        # self.tableWidget_pars.resizeRowsToContents()
        self.tableWidget_pars.setShowGrid(True)
        self.tableWidget_pars.setVerticalHeaderLabels(vertical_labels)

    def load_par(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","Table Files (*.tab);;text Files (*.txt)", options=options)
        vertical_labels = []
        if fileName:
            with open(fileName,'r') as f:
                lines = f.readlines()
                lines = [each for each in lines if not each.startswith('#')]
                how_many_pars = len(lines)
                self.tableWidget_pars.setRowCount(how_many_pars)
                self.tableWidget_pars.setColumnCount(7)
                self.tableWidget_pars.setHorizontalHeaderLabels(['Parameter','Value','Fit','Min','Max','Error','Link'])
                for i in range(len(lines)):
                    line = lines[i]
                    items = line.rstrip().rsplit('\t')
                    j = 0
                    if items[0] == '':
                        self.model.parameters.data.append([items[0],0,False,0, 0,'-',''])
                        vertical_labels.append('')
                        j += 1
                    else:
                        #add items to parameter attr
                        if len(items)==6:
                            items.append('')
                        self.model.parameters.data.append([items[0],float(items[1]),items[2]=='True',float(items[3]), float(items[4]),items[5],items[6]])
                        #add items to table view
                        if len(vertical_labels)==0:
                            vertical_labels.append('1')
                        else:
                            if vertical_labels[-1] != '':
                                vertical_labels.append('{}'.format(int(vertical_labels[-1])+1))
                            else:
                                vertical_labels.append('{}'.format(int(vertical_labels[-2])+1))
                        for item in items:
                            if j == 2:
                                check_box = QCheckBox()
                                check_box.setChecked(item=='True')
                                self.tableWidget_pars.setCellWidget(i,2,check_box)
                            else:
                                qtablewidget = QTableWidgetItem(item)
                                # qtablewidget.setTextAlignment(Qt.AlignCenter)
                                if j == 0:
                                    qtablewidget.setFont(QFont('Times',10,QFont.Bold))
                                elif j == 1:
                                    qtablewidget.setForeground(QBrush(QColor(255,0,255)))
                                self.tableWidget_pars.setItem(i,j,qtablewidget)
                            j += 1
        self.tableWidget_pars.resizeColumnsToContents()
        self.tableWidget_pars.resizeRowsToContents()
        """
        header = self.tableWidget_pars.horizontalHeader()       
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeToContents)
        """
        self.tableWidget_pars.setShowGrid(True)
        self.tableWidget_pars.setVertic
        
    def save_par(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save tab file", "", "table file (*.*)")
        with open(path,'w') as f:
            f.write(self.model.parameters.get_ascii_output())
